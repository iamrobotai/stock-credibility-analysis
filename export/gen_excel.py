# -*- coding: utf-8 -*-
"""
gen_excel.py - 分析结果 Excel 导出
多 Sheet: 概览 / D1-D8评分 / 帖子明细 / 价格分段 / 财务数据 / LLM分析
"""
import json, os, sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

BASE = os.path.dirname(os.path.abspath(__file__))
# 支持目录重组：从子目录中找到项目根
while not os.path.exists(os.path.join(BASE, "data")) and os.path.dirname(BASE) != BASE:
    BASE = os.path.dirname(BASE)
DATADIR = os.path.join(BASE, "data")
OUTDIR = os.path.join(BASE, "output")
os.makedirs(OUTDIR, exist_ok=True)

# 样式
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="2F5496")
NORMAL_FONT = Font(name="微软雅黑", size=10)
WARN_FONT = Font(name="微软雅黑", size=10, bold=True, color="C00000")
GOOD_FONT = Font(name="微软雅黑", size=10, bold=True, color="008000")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_data_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT
        cell.border = THIN_BORDER


def _auto_width(ws, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                # CJK chars count as 2
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _fmt_money(v):
    try:
        f = float(v)
        if abs(f) >= 1e8:
            return f"{f/1e8:.2f}亿"
        elif abs(f) >= 1e4:
            return f"{f/1e4:.2f}万"
        return f"{f:.2f}"
    except Exception:
        return str(v)


def _d_level(score):
    if score >= 0.7:
        return "高"
    elif score >= 0.4:
        return "中"
    return "低"


def generate(code, name="", industry=""):
    """生成 Excel 报告"""
    if not HAS_OPENPYXL:
        print("[excel] openpyxl not installed, run: pip install openpyxl")
        return None

    raw_path = os.path.join(DATADIR, f"{code}_raw.json")
    scored_path = os.path.join(DATADIR, f"{code}_scored.json")
    llm_path = os.path.join(DATADIR, f"{code}_llm.json")

    if not os.path.exists(raw_path):
        print(f"[excel] {raw_path} not found")
        return None

    raw = json.load(open(raw_path, encoding="utf-8"))
    scored = json.load(open(scored_path, encoding="utf-8")) if os.path.exists(scored_path) else {"posts": [], "segments": []}
    llm_data = json.load(open(llm_path, encoding="utf-8")) if os.path.exists(llm_path) else {}

    kline = raw.get("kline", [])
    financials = raw.get("financials", [])
    news = raw.get("news", [])
    reports = raw.get("reports", [])
    segments = scored.get("segments", [])
    posts = scored.get("posts", [])

    wb = Workbook()

    # ── Sheet 1: 概览 ──
    ws = wb.active
    ws.title = "概览"
    ws.merge_cells("A1:D1")
    ws["A1"] = f"{name}（{code}）股票可信度分析"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    row = 3
    info_rows = [
        ("股票代码", code),
        ("股票名称", name),
        ("所属行业", industry or raw.get("industry", "")),
        ("数据区间", f"{kline[0]['date']} ~ {kline[-1]['date']}" if kline else ""),
        ("交易日数", len(kline)),
        ("区间涨幅", f"{(kline[-1]['close']-kline[0]['close'])/kline[0]['close']*100:+.1f}%" if kline else ""),
        ("最高价", f"{max(b['high'] for b in kline):.2f}" if kline else ""),
        ("最低价", f"{min(b['low'] for b in kline):.2f}" if kline else ""),
        ("最新收盘", f"{kline[-1]['close']:.2f}" if kline else ""),
        ("帖子总数", len(posts)),
        ("研报数量", len(reports)),
        ("新闻数量", len(news)),
        ("广告检出", sum(1 for p in posts if p.get("D8", {}).get("is_ad"))),
        ("分段数量", len(segments)),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for label, val in info_rows:
        ws.cell(row=row, column=1, value=label).font = Font(name="微软雅黑", size=10, bold=True)
        ws.cell(row=row, column=2, value=val).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # 来源可信度均值
    row += 1
    ws.cell(row=row, column=1, value="来源类型可信度均值").font = Font(name="微软雅黑", size=11, bold=True, color="2F5496")
    row += 1
    ws.cell(row=row, column=1, value="来源类型").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).alignment = CENTER
    ws.cell(row=row, column=2, value="帖子数").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=2).alignment = CENTER
    ws.cell(row=row, column=3, value="平均可信度").font = HEADER_FONT
    ws.cell(row=row, column=3).fill = HEADER_FILL
    ws.cell(row=row, column=3).alignment = CENTER
    ws.cell(row=row, column=4, value="等级").font = HEADER_FONT
    ws.cell(row=row, column=4).fill = HEADER_FILL
    ws.cell(row=row, column=4).alignment = CENTER
    row += 1

    by_type = {}
    for p in posts:
        t = p.get("source_type", "?")
        by_type.setdefault(t, []).append(p.get("avg_d", 0))
    for t, scores in by_type.items():
        avg = sum(scores) / len(scores) if scores else 0
        level = _d_level(avg)
        ws.cell(row=row, column=1, value=t).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=len(scores)).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=round(avg, 2)).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=level).font = GOOD_FONT if avg >= 0.7 else (WARN_FONT if avg < 0.4 else NORMAL_FONT)
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).alignment = CENTER
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 10

    # ── Sheet 2: D1-D8 评分明细 ──
    ws2 = wb.create_sheet("D1-D8评分")
    headers = ["ID", "来源类型", "标题", "D1来源", "D2论据", "D3逻辑", "D4数据", "D5时效", "D6独立", "D7具体", "D8广告", "均值", "等级"]
    for c, h in enumerate(headers, 1):
        ws2.cell(row=1, column=c, value=h)
    _style_header(ws2, 1, len(headers))

    for r, p in enumerate(posts, 2):
        d8 = p.get("D8", {})
        avg = p.get("avg_d", 0)
        vals = [
            p.get("id", f"P{r-1}"),
            p.get("source_type", ""),
            str(p.get("title", ""))[:80],
            p.get("D1", 0),
            p.get("D2", 0),
            p.get("D3", 0),
            p.get("D4", 0),
            p.get("D5", 0),
            p.get("D6", 0),
            p.get("D7", 0),
            "⚠️" if d8.get("is_ad") else "—",
            round(avg, 2),
            _d_level(avg),
        ]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=r, column=c, value=v)
        _style_data_row(ws2, r, len(headers))
        # 广告行标红
        if d8.get("is_ad"):
            for c in range(1, len(headers) + 1):
                ws2.cell(row=r, column=c).fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")

    _auto_width(ws2)
    ws2.freeze_panes = "A2"

    # ── Sheet 3: 价格分段 ──
    ws3 = wb.create_sheet("价格分段")
    seg_headers = ["波段ID", "方向", "起始日", "结束日", "起始价", "结束价", "涨跌幅%", "交易日数"]
    for c, h in enumerate(seg_headers, 1):
        ws3.cell(row=1, column=c, value=h)
    _style_header(ws3, 1, len(seg_headers))

    for r, s in enumerate(segments, 2):
        vals = [s.get("id", ""), s.get("direction", ""), s.get("start_date", ""),
                s.get("end_date", ""), s.get("start_price", 0), s.get("end_price", 0),
                s.get("pct", 0), s.get("bars", 0)]
        for c, v in enumerate(vals, 1):
            ws3.cell(row=r, column=c, value=v)
        _style_data_row(ws3, r, len(seg_headers))
        # 涨红跌绿
        pct_cell = ws3.cell(row=r, column=7)
        if s.get("direction") == "涨":
            pct_cell.font = Font(name="微软雅黑", size=10, bold=True, color="C00000")
        else:
            pct_cell.font = Font(name="微软雅黑", size=10, bold=True, color="008000")

    _auto_width(ws3)
    ws3.freeze_panes = "A2"

    # ── Sheet 4: 研报 ──
    if reports:
        ws4 = wb.create_sheet("研报")
        rep_headers = ["机构", "评级", "EPS2026", "PE2026", "报告标题", "日期", "原始链接"]
        for c, h in enumerate(rep_headers, 1):
            ws4.cell(row=1, column=c, value=h)
        _style_header(ws4, 1, len(rep_headers))
        for r, rep in enumerate(reports, 2):
            vals = [rep.get("org", ""), rep.get("rating", ""), rep.get("eps_2026", ""),
                    rep.get("pe_2026", ""), rep.get("title", ""), rep.get("date", ""),
                    rep.get("url", "")]
            for c, v in enumerate(vals, 1):
                ws4.cell(row=r, column=c, value=v)
            _style_data_row(ws4, r, len(rep_headers))
        _auto_width(ws4)
        ws4.freeze_panes = "A2"

    # ── Sheet 5: 财务数据 ──
    if financials:
        ws5 = wb.create_sheet("财务数据")
        # 财务数据是字典列表，每个字典有 "指标" 键
        all_keys = set()
        for f in financials:
            all_keys.update(f.keys())
        all_keys = sorted(all_keys)
        for c, k in enumerate(all_keys, 1):
            ws5.cell(row=1, column=c, value=k)
        _style_header(ws5, 1, len(all_keys))
        for r, f in enumerate(financials, 2):
            for c, k in enumerate(all_keys, 1):
                ws5.cell(row=r, column=c, value=f.get(k, ""))
            _style_data_row(ws5, r, len(all_keys))
        _auto_width(ws5)
        ws5.freeze_panes = "A2"

    # ── Sheet 6: LLM 分析 ──
    if llm_data.get("success") and llm_data.get("analysis"):
        ws6 = wb.create_sheet("LLM分析")
        a = llm_data["analysis"]
        stats = llm_data.get("stats", {})

        ws6.merge_cells("A1:B1")
        ws6["A1"] = f"LLM 深度分析 (qwen3:4b)"
        ws6["A1"].font = TITLE_FONT

        row = 3
        llm_rows = [
            ("模型", "qwen3:4b"),
            ("耗时", f"{stats.get('duration_s', 0):.1f}s"),
            ("Tokens", stats.get("tokens", 0)),
            ("速度", f"{stats.get('tps', 0):.0f} tok/s"),
            ("", ""),
            ("总体评价", a.get("overall_assessment", "")),
            ("", ""),
        ]
        for label, val in llm_rows:
            if label:
                ws6.cell(row=row, column=1, value=label).font = Font(name="微软雅黑", size=10, bold=True)
                ws6.cell(row=row, column=2, value=val).font = NORMAL_FONT
                ws6.cell(row=row, column=1).border = THIN_BORDER
                ws6.cell(row=row, column=2).border = THIN_BORDER
            row += 1

        if a.get("key_risks"):
            ws6.cell(row=row, column=1, value="关键风险").font = Font(name="微软雅黑", size=11, bold=True, color="C00000")
            row += 1
            for risk in a["key_risks"]:
                ws6.cell(row=row, column=1, value="• " + risk).font = NORMAL_FONT
                ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                row += 1

        if a.get("catalysts"):
            row += 1
            ws6.cell(row=row, column=1, value="催化剂").font = Font(name="微软雅黑", size=11, bold=True, color="008000")
            row += 1
            for cat in a["catalysts"]:
                ws6.cell(row=row, column=1, value="• " + cat).font = NORMAL_FONT
                ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                row += 1

        if a.get("ad_warning"):
            row += 1
            ws6.cell(row=row, column=1, value="广告甄别").font = Font(name="微软雅黑", size=11, bold=True, color="B8860B")
            row += 1
            ws6.cell(row=row, column=1, value=a["ad_warning"]).font = NORMAL_FONT
            ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

        ws6.column_dimensions["A"].width = 20
        ws6.column_dimensions["B"].width = 80

    # ── Sheet 7: 新闻 ──
    if news:
        ws7 = wb.create_sheet("新闻")
        news_headers = ["日期", "来源", "标题", "原始链接"]
        for c, h in enumerate(news_headers, 1):
            ws7.cell(row=1, column=c, value=h)
        _style_header(ws7, 1, len(news_headers))
        for r, n in enumerate(news, 2):
            vals = [str(n.get("time", ""))[:10], n.get("source", ""), n.get("title", ""), n.get("url", "")]
            for c, v in enumerate(vals, 1):
                ws7.cell(row=r, column=c, value=v)
            _style_data_row(ws7, r, len(news_headers))
        _auto_width(ws7)
        ws7.freeze_panes = "A2"

    # 保存
    safe_name = name.replace("/", "_").replace(" ", "") if name else code
    outname = f"{safe_name}_{code}_分析.xlsx"
    outpath = os.path.join(OUTDIR, outname)
    wb.save(outpath)
    print(f"[excel] {outpath} ({os.path.getsize(outpath)//1024}KB)")
    return outpath


if __name__ == "__main__":
    code = sys.argv[1] if len(sys.argv) > 1 else "002371"
    name = sys.argv[2] if len(sys.argv) > 2 else ""
    industry = sys.argv[3] if len(sys.argv) > 3 else ""
    generate(code, name, industry)
